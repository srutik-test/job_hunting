"""
Excel / CSV import & export.

Export columns follow the evidence-first contract: Company Name, Website,
Location, HR Name, Designation, HR Email, LinkedIn Profile, Source,
Source URL, Discovery Method, Verification Status, Confidence, Date Found.

Contacts without email evidence are clearly marked; unverified addresses are
never presented as verified.
"""

import io
import uuid
from datetime import datetime, timezone
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.domain import CompanyInput

EXPORT_HEADERS = [
    "Company Name",
    "HR Mails",
    "Phone Number",
    "Company Website",
    "LinkedIn URL",
    "Location",
]

_HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def build_results_workbook(rows: List[dict]) -> bytes:
    """rows: dicts with keys matching the internal representation."""
    wb = Workbook()
    ws = wb.active
    ws.title = "HR Contacts"

    for col, title in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r, row in enumerate(rows, start=2):
        email = row.get("email")
        phone = row.get("phone")
        ws.cell(row=r, column=1, value=row.get("company_name", ""))
        ws.cell(row=r, column=2, value=email if email else "No email evidence")
        ws.cell(row=r, column=3, value=phone if phone else "-")
        ws.cell(row=r, column=4, value=row.get("website", ""))
        ws.cell(row=r, column=5, value=row.get("linkedin_url") or "-")
        ws.cell(row=r, column=6, value=row.get("location", ""))

    for idx, header in enumerate(EXPORT_HEADERS, start=1):
        width = max(len(header) + 6, 20)
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



# --------------------------------------------------------------- import parsing
_COLUMN_ALIASES = {
    "name": {"company name", "company", "organization", "organisation", "name"},
    "website": {"website", "url", "site", "domain", "company website", "web"},
    "location": {"location", "city", "address", "hq", "headquarters", "state"},
    "linkedin_url": {
        "linkedin url",
        "linkedin",
        "linkedin profile",
        "linkedin page",
        "li url",
    },
    "industry": {"industry", "sector", "vertical"},
}


def _match_column(header: str) -> str | None:
    clean = " ".join(str(header).strip().lower().split())
    for field, aliases in _COLUMN_ALIASES.items():
        if clean in aliases or any(clean.startswith(a) for a in aliases):
            return field
    return None


def canonical_company_key(name: str, website: str) -> tuple[str, str]:
    from app.services.orchestrator import normalize_website, domain_of

    clean_name = " ".join(str(name or "").strip().lower().split())
    norm_url = normalize_website(website) or str(website or "").strip().lower()
    dom = domain_of(norm_url)
    return (clean_name, dom)


def parse_companies_from_file(content: bytes, filename: str) -> List[CompanyInput]:
    """Parse an xlsx/csv upload into validated, deduplicated CompanyInput rows."""
    import pandas as pd
    from app.services.orchestrator import normalize_website

    name_l = (filename or "").lower()
    try:
        if name_l.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
    except Exception as exc:
        raise ValueError(f"Could not parse file: {exc}")

    rename: dict[str, str] = {}
    for col in df.columns:
        match = _match_column(col)
        if match and match not in rename.values():
            rename[col] = match

    df = df.rename(columns=rename)
    if "name" not in df.columns or "website" not in df.columns:
        raise ValueError(
            "File must contain 'Company Name' and 'Website' columns "
            f"(detected columns: {', '.join(map(str, df.columns))})."
        )

    out: List[CompanyInput] = []
    seen_keys: set[tuple[str, str]] = set()

    for _, row in df.iterrows():
        name = str(row.get("name", "")).strip()
        website = str(row.get("website", "")).strip()
        if not name or name.lower() == "nan" or not website or website.lower() == "nan":
            continue

        norm_website = normalize_website(website) or website.strip().rstrip("/")
        key = canonical_company_key(name, norm_website)
        if key in seen_keys:
            continue  # Keep the first occurrence only
        seen_keys.add(key)

        out.append(
            CompanyInput(
                name=name[:255],
                website=norm_website[:1024],
                location=(
                    str(row.get("location", "")).strip()
                    if str(row.get("location", "")) != "nan"
                    else ""
                )[:255],
                linkedin_url=(
                    str(row.get("linkedin_url", "")).strip()
                    if str(row.get("linkedin_url", "")) != "nan"
                    else ""
                )[:1024],
                industry=(
                    str(row.get("industry", "")).strip()
                    if str(row.get("industry", "")) != "nan"
                    else ""
                )[:255],
            )
        )
    if not out:
        raise ValueError("No valid company rows found in the uploaded file.")
    return out[:100]


def sample_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"
    headers = ["Company Name", "Location", "Website", "LinkedIn URL", "Industry"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    samples = [
        (
            "Example Software Ltd",
            "London",
            "https://example-software.co.uk",
            "https://linkedin.com/company/example-software",
            "IT Services",
        ),
    ]
    for r, s in enumerate(samples, 2):
        for c, v in enumerate(s, 1):
            ws.cell(row=r, column=c, value=v)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(h) + 6, 18)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def new_export_filename(prefix: str = "hr_contacts") -> str:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return f"{prefix}_{stamp}_{uuid.uuid4().hex[:6]}.xlsx"
