"""
Excel and CSV Exporter module.
Generates structured Excel (.xlsx) and CSV exports with styling,
verification statuses, and confidence ratings.
"""

import io
from typing import List, Dict, Any
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd


class ExcelExporter:
    """Exports extraction results into formatted Excel and CSV files."""

    COLUMNS = [
        ("company_name", "Company Name", 25),
        ("location", "Location", 18),
        ("website", "Website", 30),
        ("linkedin_url", "LinkedIn", 35),
        ("hr_email", "HR Email", 28),
        ("recruitment_email", "Recruitment Email", 28),
        ("careers_email", "Careers Email", 28),
        ("general_email", "General Email", 28),
        ("hr_name", "HR Name", 22),
        ("hr_position", "HR Position", 26),
        ("linkedin_profile", "LinkedIn Profile", 38),
        ("source", "Source", 35),
        ("confidence_score", "Confidence Score", 18),
        ("status", "Verification Status", 25),
        ("created_at", "Extraction Date", 20),
    ]

    @classmethod
    def export_to_excel(cls, results: List[Dict[str, Any]]) -> bytes:
        """Create a styled Excel (.xlsx) workbook."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HR Contacts Intelligence"

        # Headers
        headers = [title for _, title, _ in cls.COLUMNS]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(
            start_color="0F172A", end_color="0F172A", fill_type="solid"
        )
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(
            horizontal="center", vertical="center", wrap_text=False
        )

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Add data rows
        row_alt_fill = PatternFill(
            start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"
        )
        regular_font = Font(name="Segoe UI", size=10)
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        for row_idx, r in enumerate(results, start=2):
            row_data = []
            for key, _, _ in cls.COLUMNS:
                val = r.get(key, "")
                if key == "confidence_score":
                    val = f"{val}%" if isinstance(val, (int, float)) else str(val)
                elif key == "created_at" and val:
                    try:
                        # Format ISO date
                        val = str(val)[:19].replace("T", " ")
                    except Exception:
                        pass
                row_data.append(val if val is not None else "Not Publicly Available")

            ws.append(row_data)

            # Apply row styling
            is_alt = row_idx % 2 == 0
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                if is_alt:
                    cell.fill = row_alt_fill

        # Auto-adjust column widths
        for col_idx, (_, _, default_width) in enumerate(cls.COLUMNS, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = default_width

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    @classmethod
    def export_to_csv(cls, results: List[Dict[str, Any]]) -> str:
        """Create a clean CSV string."""
        data_rows = []
        for r in results:
            row = {}
            for key, title, _ in cls.COLUMNS:
                val = r.get(key, "")
                if key == "confidence_score":
                    val = f"{val}%" if isinstance(val, (int, float)) else str(val)
                row[title] = val if val is not None else "Not Publicly Available"
            data_rows.append(row)

        df = pd.DataFrame(data_rows)
        return df.to_csv(index=False)
