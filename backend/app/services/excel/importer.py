"""
Excel and CSV Importer module.
Supports robust parsing, fuzzy column mapping, validation, and sample template generation.
"""

import io
import re
from typing import List, Dict, Any, Tuple, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.schemas.company import CompanyBase, CompanyBatchUploadPreview


class ExcelImporter:
    """Parser and validator for company list imports."""

    # Fuzzy column aliases
    COLUMN_ALIASES = {
        "name": [
            "company name",
            "company",
            "company_name",
            "organization",
            "business name",
            "firm",
        ],
        "location": [
            "location",
            "city",
            "headquarters",
            "hq",
            "company location",
            "country",
            "address",
        ],
        "website": [
            "website",
            "company website",
            "url",
            "web",
            "domain",
            "site",
            "homepage",
        ],
        "linkedin_url": [
            "linkedin",
            "linkedin url",
            "company linkedin",
            "linkedin_url",
            "linkedin link",
            "company linkedin url",
        ],
    }

    @classmethod
    def parse_file(
        cls, file_content: bytes, filename: str
    ) -> Tuple[List[CompanyBase], CompanyBatchUploadPreview]:
        """Parse Excel or CSV bytes into validated Company records and preview metadata."""
        if filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(file_content))
        elif filename.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(file_content))
        else:
            raise ValueError(
                "Unsupported file format. Please upload .xlsx, .xls, or .csv"
            )

        # Normalize column headers
        raw_columns = [str(c).strip() for c in df.columns]
        mapped_columns: Dict[str, str] = {}  # target_key -> raw_column_name

        for raw_col in raw_columns:
            cleaned_col = raw_col.lower().replace("_", " ").replace("-", " ")
            for target_field, aliases in cls.COLUMN_ALIASES.items():
                if cleaned_col in aliases or any(
                    alias in cleaned_col for alias in aliases
                ):
                    if target_field not in mapped_columns:
                        mapped_columns[target_field] = raw_col

        # Check required fields
        warnings: List[str] = []
        missing_columns: List[str] = []
        if "name" not in mapped_columns:
            missing_columns.append("Company Name")
        if "website" not in mapped_columns:
            missing_columns.append("Website")

        if missing_columns:
            raise ValueError(
                f"Missing required columns in file: {', '.join(missing_columns)}. Found columns: {', '.join(raw_columns)}"
            )

        valid_companies: List[CompanyBase] = []
        invalid_count = 0

        for idx, row in df.iterrows():
            try:
                name_val = (
                    str(row[mapped_columns["name"]]).strip()
                    if pd.notna(row[mapped_columns["name"]])
                    else ""
                )
                website_val = (
                    str(row[mapped_columns["website"]]).strip()
                    if pd.notna(row[mapped_columns["website"]])
                    else ""
                )

                location_val = ""
                if "location" in mapped_columns and pd.notna(
                    row[mapped_columns["location"]]
                ):
                    location_val = str(row[mapped_columns["location"]]).strip()

                linkedin_val = ""
                if "linkedin_url" in mapped_columns and pd.notna(
                    row[mapped_columns["linkedin_url"]]
                ):
                    linkedin_val = str(row[mapped_columns["linkedin_url"]]).strip()

                if (
                    not name_val
                    or name_val.lower() == "nan"
                    or not website_val
                    or website_val.lower() == "nan"
                ):
                    invalid_count += 1
                    continue

                company = CompanyBase(
                    name=name_val,
                    location=location_val if location_val.lower() != "nan" else "",
                    website=website_val,
                    linkedin_url=linkedin_val if linkedin_val.lower() != "nan" else "",
                )
                valid_companies.append(company)
            except Exception as e:
                invalid_count += 1
                warnings.append(f"Row {idx + 2}: {str(e)}")

        preview = CompanyBatchUploadPreview(
            total_parsed=len(df),
            valid_count=len(valid_companies),
            invalid_count=invalid_count,
            preview_items=valid_companies[:10],
            detected_columns=raw_columns,
            missing_columns=missing_columns,
            warnings=warnings[:10],
        )

        return valid_companies, preview

    @classmethod
    def generate_sample_excel(cls) -> bytes:
        """Generate a professionally styled sample Excel template."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Companies Template"

        headers = ["Company Name", "Location", "Website", "LinkedIn URL"]
        ws.append(headers)

        sample_rows = [
            [
                "Aspire Softserv",
                "Ahmedabad",
                "https://aspiresoftserv.com",
                "https://linkedin.com/company/aspiresoftserv",
            ],
            [
                "Simform",
                "Ahmedabad",
                "https://simform.com",
                "https://linkedin.com/company/simform",
            ],
            [
                "Bacancy Technology",
                "Ahmedabad",
                "https://bacancytechnology.com",
                "https://linkedin.com/company/bacancy-technology",
            ],
            [
                "Radixweb",
                "Ahmedabad",
                "https://radixweb.com",
                "https://linkedin.com/company/radixweb",
            ],
            [
                "TatvaSoft",
                "Ahmedabad",
                "https://tatvasoft.com",
                "https://linkedin.com/company/tatvasoft",
            ],
        ]

        for r in sample_rows:
            ws.append(r)

        # Style header
        header_fill = PatternFill(
            start_color="1E293B", end_color="1E293B", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    @classmethod
    def generate_sample_csv(cls) -> str:
        """Generate sample CSV template string."""
        return (
            "Company Name,Location,Website,LinkedIn URL\n"
            "Aspire Softserv,Ahmedabad,https://aspiresoftserv.com,https://linkedin.com/company/aspiresoftserv\n"
            "Simform,Ahmedabad,https://simform.com,https://linkedin.com/company/simform\n"
            "Bacancy Technology,Ahmedabad,https://bacancytechnology.com,https://linkedin.com/company/bacancy-technology\n"
            "Radixweb,Ahmedabad,https://radixweb.com,https://linkedin.com/company/radixweb\n"
            "TatvaSoft,Ahmedabad,https://tatvasoft.com,https://linkedin.com/company/tatvasoft\n"
        )
